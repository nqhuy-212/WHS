import sys
import os
import time
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Generator
from dotenv import load_dotenv
from decimal import Decimal

# PyQt5
from PyQt5.QtCore import QDate
from PyQt5.QtGui import QColor, QPixmap
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QMessageBox, QFileDialog, QTableWidgetItem
)
from PyQt5.QtMultimedia import QSound
from PyQt5.uic import loadUiType
from PyQt5.QtWebEngineWidgets import QWebEngineView

# SQL & Database
import pyodbc
from sqlalchemy import create_engine, VARCHAR, NVARCHAR, INTEGER, DATE, DECIMAL
from sqlalchemy.engine import URL, Engine
from sqlalchemy.orm import sessionmaker, declarative_base

# Excel & Pandas
import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# PDF & QR Code
import qrcode
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4

# Windows API (chỉ dùng khi chạy trên Windows)
import win32print
import win32api

# Load UI
import qdarkstyle
import resources_rc

# Đăng ký font Unicode cho ReportLab
pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
pdfmetrics.registerFont(TTFont('Arial-Bold', 'arialbd.ttf'))

def get_resource_path(relative_path):
    """Trả về đường dẫn đầy đủ đến tài nguyên."""
    if getattr(sys, 'frozen', False):  # Kiểm tra nếu đang chạy file .exe
        base_path = sys._MEIPASS
    else:  # Nếu đang chạy bằng Python gốc
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def connect_to_db(): 
    BASE_DIR = Path(__file__).resolve().parent
    env_file = get_resource_path(".env")
    load_dotenv(env_file)
    try:
        connection = pyodbc.connect(
        'DRIVER={SQL Server};'
        f'SERVER={os.getenv("SERVER")};'
        f'DATABASE={os.getenv("DB")};'
        f'UID={os.getenv("UID")};'
        f'PWD={os.getenv("PASSWORD")}'
    )
        # connection = sqlite3.connect(db_file)
        return connection
    except pyodbc.Error as e:
        print(f"Lỗi khi kết nối tới máy chủ: {e}")
        return None

def table_to_dataframe(table_widget,headers):
        rows = table_widget.rowCount()
        columns = table_widget.columnCount()
        
        # Lấy tiêu đề cột
        # headers = [table_widget.horizontalHeaderItem(i).text() for i in range(columns)]
        
        # Lấy dữ liệu từ bảng
        data = []
        for row in range(rows):
            row_data = []
            for column in range(columns):
                item = table_widget.item(row, column)
                row_data.append(item.text() if item else '')  # Lấy text từ ô, nếu không có thì gán chuỗi rỗng
            data.append(row_data)
        
        # Tạo DataFrame
        df = pd.DataFrame(data, columns=headers)
        return df
    
ui, _ = loadUiType(get_resource_path('Fb_Whs.ui'))

#config URL cho engine
BASE_DIR = Path(__file__).resolve().parent
env_file = get_resource_path(".env")
load_dotenv(env_file)

class Settings():
    API_PREFIX = ''
    DATABASE_1_URL = URL.create(
        "mssql+pyodbc",
        username=os.getenv("UID"),
        password=os.getenv("PASSWORD"),
        host=os.getenv("SERVER"),
        port=1433,
        database=os.getenv("DB"),
        query={
           "driver": "ODBC Driver 17 for SQL Server",
           "TrustServerCertificate": "yes" 
        }
    )

settings = Settings()
#tạo engine để kêt nối database
engine_1 = create_engine(settings.DATABASE_1_URL, pool_pre_ping=True)
SessionLocal_1 = sessionmaker(autocommit=False, autoflush=False, bind=engine_1)

Base = declarative_base()

def get_db_1() -> Generator:
    try:
        db = SessionLocal_1()
        yield db
    finally:
        db.close()
#hàm import to sql       
def import_to_sql(df: DataFrame, table_name: str, dtype: dict, engine: Engine):
    # Show processing message
    processing_message = QMessageBox()
    processing_message.setWindowTitle("Đang xử lý")
    processing_message.setText("Đang xử lý dữ liệu, vui lòng chờ...")
    # processing_message.setStandardButtons(QMessageBox.NoButton)
    processing_message.setModal(True)
    processing_message.show()
    QApplication.processEvents()  # Ensure UI updates during processing
    time.sleep(0.01)  # Simulate processing time
    try:
        with engine.connect() as connection:
            df.to_sql(name=table_name, con=connection, if_exists="append", index=False, dtype=dtype)
            
        processing_message.close()
    except Exception as e:
        processing_message.close()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

class MainApp(QMainWindow,ui):
    def __init__(self):
        QMainWindow.__init__(self)
        self.setupUi(self)
        
        self.tabWidget.setCurrentIndex(0)
        self.tabWidget.tabBar().setVisible(False)
        self.menuBar.setVisible(False)
        
        # self.toolBar.setVisible(False)
        self.bt001.clicked.connect(self.login)
        self.menu11.triggered.connect(self.show_nhap_kho_tab)
        self.menu13.triggered.connect(self.show_chuyen_vi_tri_tab)
        self.menu21.triggered.connect(self.show_xa_vai_tab)
        self.menu22.triggered.connect(self.show_xuat_kho_tab)
        self.menu31.triggered.connect(self.show_ton_kho_tab)
        # self.menu41.triggered.connect(self.show_bao_cao_tab)
        self.menu51.triggered.connect(self.show_login_tab)
        # self.act1.triggered.connect(self.show_nhap_kho_tab)
        # self.act2.triggered.connect(self.show_xuat_kho_tab)
        # self.act3.triggered.connect(self.show_ton_kho_tab)
        # self.act4.triggered.connect(self.show_login_tab)
        self.de101.setDate(QDate.currentDate().addDays(-90))
        self.de102.setDate(QDate.currentDate())
        self.de201.setDate(QDate.currentDate().addDays(-90))
        self.de202.setDate(QDate.currentDate())
        self.de301.setDate(QDate.currentDate().addDays(-90))
        self.de302.setDate(QDate.currentDate())
        # self.tb107.setValidator(QIntValidator(0, 999)) 
        # self.tb108.setValidator(QDoubleValidator(0.0, 1000.0, 2))
        ####
        self.bt101.clicked.connect(self.print_labels_3x2)
        self.bt102.clicked.connect(self.delete_selected_rows)
        self.bt103.clicked.connect(self.import_from_excel)
        self.bt203.clicked.connect(self.show_QR_xa_vai_tab)
        self.bt303.clicked.connect(self.show_QR_xuat_kho_tab)
        self.bt403.clicked.connect(self.show_so_do_kho_tab)
        self.bt104.clicked.connect(self.tai_xuong_file_mau)
        self.bt204.clicked.connect(self.tai_xuong_file_xa_vai)
        self.bt304.clicked.connect(self.tai_xuong_file_xuat_kho)
        self.bt404.clicked.connect(self.tai_xuong_file_ton_kho)
        self.bt105.clicked.connect(self.search_nhap_kho)
        self.bt205.clicked.connect(self.search_xa_vai)
        self.bt305.clicked.connect(self.search_xuat_kho)
        self.bt405.clicked.connect(self.search_ton_kho)
        self.bt501.clicked.connect(self.show_xa_vai_tab)
        self.bt601.clicked.connect(self.show_xuat_kho_tab)
        self.bt801.clicked.connect(self.show_ton_kho_tab)
        self.bt901.clicked.connect(self.show_ton_kho_tab)
        self.tb501.returnPressed.connect(self.handle_scan_xa_vai)  # Xử lý khi nhấn Enter
        self.tb601.returnPressed.connect(self.handle_scan_xuat_kho)  # Xử lý khi nhấn Enter
        self.tb701.returnPressed.connect(self.handle_scan_chuyen_vi_tri)  # Xử lý khi nhấn Enter
        ####
        # self.tableWidget.cellChanged.connect(self.cap_nhat_QR)

    def login(self):
        fty = self.cb001.currentText()
        un = self.tb001.text()
        pw = self.tb002.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        cursor.execute("""
                SELECT macongty, masothe, hoten, phongban, Kho 
                FROM HR.DBO.NHANVIEN 
                WHERE macongty = ? AND masothe = ? AND matkhau = ? AND Kho IS NOT NULL
            """, (fty, un, pw))
        result = cursor.fetchone()
             
        if result:
            self.menuBar.setVisible(True)
            self.tabWidget.setCurrentIndex(1)
            self.lb003.setText("Phần mềm quản lý kho vải")
            self.lb000.setText(result[0])
            self.lb001.setText(result[2])
            QSound.play(":/sounds/sounds/success.wav") # Phát âm thanh thành công
            #progress bar
            self.progressBar.setValue(0)  # Khởi tạo giá trị là 0
            self.progressBar.setMinimum(0)
            self.progressBar.setMaximum(100)
            self.progressBar_2.setValue(0)  # Khởi tạo giá trị là 0
            self.progressBar_2.setMinimum(0)
            self.progressBar_2.setMaximum(100)
            self.progressBar_5.setValue(0)  # Khởi tạo giá trị là 0
            self.progressBar_5.setMinimum(0)
            self.progressBar_5.setMaximum(100)
            self.progressBar_6.setValue(0)  # Khởi tạo giá trị là 0
            self.progressBar_6.setMinimum(0)
            self.progressBar_6.setMaximum(100)
        else:
            QSound.play(":/sounds/sounds/error.wav") # Phát âm thanh lỗi
            self.lb002.setText("Tài khoản hoặc mật khẩu không đúng!")
    def search_nhap_kho(self):
        tu_ngay = self.de101.date().toString("yyyy-MM-dd")
        den_ngay = self.de102.date().toString("yyyy-MM-dd")
        style = self.tb101.text()
        mo = self.tb102.text()
        lot = self.tb103.text()
        mau = self.tb104.text()
        vitri = self.tb105.text()
        nha_may = self.lb000.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        cursor.execute(f"""
                SELECT NGAY_NHAN,STYLE,MO,LOAI_VAI,DVT,LOT,MAU,CUON_SO,SO_YARD,VI_TRI,ID,
                CONVERT(VARCHAR, THOI_GIAN_XA, 120) AS THOI_GIAN_XA,
                CONVERT(VARCHAR, THOI_GIAN_XUAT_KHO, 120) AS THOI_GIAN_XUAT_KHO,
                TRANG_THAI
                FROM DANH_SACH_CUON_VAI
                WHERE NGAY_NHAN BETWEEN '{tu_ngay}' AND '{den_ngay}' 
                AND STYLE LIKE '%{style}%'
                AND MO LIKE '%{mo}%'
                AND LOT LIKE '%{lot}%'
                AND MAU LIKE '%{mau}%'
                AND ISNULL(VI_TRI,'') LIKE '%{vitri}%'
                AND NHA_MAY LIKE '%{nha_may}%'
            """)
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget.setItem(row_idx, col_idx, item)
            
            # Cập nhật progress bar
            self.progressBar.setValue(int((row_idx + 1) * 100 / len(results)))
        
        # Gọi hàm tổng số dòng
        self.tong_so_dong()
        
    def search_xa_vai(self):
        tu_ngay = self.de201.date().toString("yyyy-MM-dd")
        den_ngay = self.de202.date().toString("yyyy-MM-dd")
        style = self.tb201.text()
        mo = self.tb202.text()
        lot = self.tb203.text()
        mau = self.tb204.text()
        vitri = self.tb205.text()
        nha_may = self.lb000.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        sql = f"""
                SELECT NGAY_NHAN,STYLE,MO,LOAI_VAI,DVT,LOT,MAU,CUON_SO,SO_YARD,VI_TRI,ID,
                CONVERT(VARCHAR, THOI_GIAN_XA, 120) AS THOI_GIAN_XA,
                CAST(DATEDIFF(MINUTE,THOI_GIAN_XA,GETDATE())/60.0 AS DEC(10,2)) AS SO_GIO
                FROM DANH_SACH_CUON_VAI
                WHERE (CAST(THOI_GIAN_XA AS DATE) BETWEEN '{tu_ngay}' AND '{den_ngay}')
                AND STYLE LIKE '%{style}%'
                AND MO LIKE '%{mo}%'
                AND LOT LIKE '%{lot}%'
                AND MAU LIKE '%{mau}%'
                AND ISNULL(VI_TRI,'') LIKE '%{vitri}%'
                AND NHA_MAY LIKE '%{nha_may}%'
                AND TRANG_THAI = N'Xả vải'
            """
        cursor.execute(sql)
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget_2.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget_2.insertRow(row_idx)
            so_gio = row_data[12]
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget_2.setItem(row_idx, col_idx, item)
                if so_gio >= 24:
                    item.setBackground(QColor("#008a10"))          
            # Cập nhật progress bar
            self.progressBar_2.setValue(int((row_idx + 1) * 100 / len(results)))
        
        # Gọi hàm tổng số dòng
        self.tong_so_dong_xa_vai()
    
    def search_xuat_kho(self):
        tu_ngay = self.de301.date().toString("yyyy-MM-dd")
        den_ngay = self.de302.date().toString("yyyy-MM-dd")
        style = self.tb301.text()
        mo = self.tb302.text()
        lot = self.tb303.text()
        mau = self.tb304.text()
        nha_may = self.lb000.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        sql = f"""
                SELECT NGAY_NHAN,STYLE,MO,LOAI_VAI,DVT,LOT,MAU,CUON_SO,SO_YARD,VI_TRI,ID,
                CONVERT(VARCHAR, THOI_GIAN_XA, 120) AS THOI_GIAN_XA,
                CONVERT(VARCHAR, THOI_GIAN_XUAT_KHO, 120) AS THOI_GIAN_XUAT_KHO 
                FROM DANH_SACH_CUON_VAI
                WHERE (CAST(THOI_GIAN_XUAT_KHO AS DATE) BETWEEN '{tu_ngay}' AND '{den_ngay}')
                AND STYLE LIKE '%{style}%'
                AND MO LIKE '%{mo}%'
                AND LOT LIKE '%{lot}%'
                AND MAU LIKE '%{mau}%'
                AND NHA_MAY LIKE '%{nha_may}%'
                AND TRANG_THAI = N'Xuất kho'
            """
        cursor.execute(sql)
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget_5.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        for row_idx, row_data in enumerate(results):
            self.tableWidget_5.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget_5.setItem(row_idx, col_idx, item)
            
            # Cập nhật progress bar
            self.progressBar_5.setValue(int((row_idx + 1) * 100 / len(results)))
        
        # Gọi hàm tổng số dòng
        self.tong_so_dong_xuat_kho()   
    
    def search_ton_kho(self):
        style = self.tb401.text()
        mo = self.tb402.text()
        lot = self.tb403.text()
        mau = self.tb404.text()
        vitri = self.tb405.text()
        nha_may = self.lb000.text()
        
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        sql = f"""
                SELECT NGAY_NHAN,STYLE,MO,LOAI_VAI,DVT,LOT,MAU,CUON_SO,SO_YARD,VI_TRI,ID,
                CONVERT(VARCHAR, THOI_GIAN_XA, 120) AS THOI_GIAN_XA,
                CONVERT(VARCHAR, THOI_GIAN_XUAT_KHO, 120) AS THOI_GIAN_XUAT_KHO 
                FROM DANH_SACH_CUON_VAI
                WHERE STYLE LIKE '%{style}%'
                AND MO LIKE '%{mo}%'
                AND LOT LIKE '%{lot}%'
                AND MAU LIKE '%{mau}%'
                AND ISNULL(VI_TRI,'') LIKE '%{vitri}%'
                AND NHA_MAY LIKE '%{nha_may}%'
                AND TRANG_THAI = N'Nhập kho'
            """
        cursor.execute(sql)
        results = cursor.fetchall()
        # Xóa dữ liệu cũ trong TableWidget
        self.tableWidget_6.setRowCount(0)
        
        # Hiển thị dữ liệu trong TableWidget
        tong_so_yards = Decimal('0')
        for row_idx, row_data in enumerate(results):
            self.tableWidget_6.insertRow(row_idx)
            for col_idx, value in enumerate(row_data):
                item = QTableWidgetItem(str(value) if value is not None else "")
                self.tableWidget_6.setItem(row_idx, col_idx, item)
                if col_idx == 8:
                    tong_so_yards += Decimal(value)  # Cộng giá trị vào tổng
            
            # Cập nhật progress bar
            self.progressBar_6.setValue(int((row_idx + 1) * 100 / len(results)))
        
        # Gọi hàm tổng số dòng
        self.tong_so_dong_ton_kho()   
        self.lb44.setText(f"{tong_so_yards:,.0f}")
        
    def show_login_tab(self):
        self.tabWidget.setCurrentIndex(0) 
        self.tb001.setText("")
        self.tb002.setText("")
        self.lb000.setText("")
        self.lb001.setText("")
        self.lb002.setText("")
        self.lb003.setText("Phần mềm quản lý kho vải")
        self.menuBar.setVisible(False)
                        
    def show_nhap_kho_tab(self):
        self.tabWidget.setCurrentIndex(1) 
        self.lb003.setText("Nhập kho")
     
    def show_xa_vai_tab(self):
        self.menuBar.setVisible(True)
        self.tabWidget.setCurrentIndex(2) 
        self.lb003.setText("Xả vải")
     
    def show_QR_xa_vai_tab(self):
        self.menuBar.setVisible(False)
        self.tabWidget.setCurrentIndex(5) 
        self.lb003.setText("Quét mã QR xả vải")
        self.lb501.setText("")
        self.lb502.setText("")
        self.lb503.setText("")
        self.lb504.setText("")
        self.tb501.setFocus()
    
    def show_QR_xuat_kho_tab(self):
        self.menuBar.setVisible(False)
        self.tabWidget.setCurrentIndex(6) 
        self.lb003.setText("Quét mã QR xuất kho")
        self.lb601.setText("")
        self.lb602.setText("")
        self.lb603.setText("")
        self.lb604.setText("")
        self.tb601.setFocus()
    
    def show_so_do_kho_tab(self):
        self.menuBar.setVisible(False)
        nha_may = self.lb000.text()
        if nha_may == 'NT1':
            self.tabWidget.setCurrentIndex(8) 
            self.lb003.setText("Sơ đồ kho vải Nam Thuận 1")

        else:
            self.tabWidget.setCurrentIndex(9) 
            self.lb003.setText("Sơ đồ kho vải Nam Thuận 2")
                   
    def show_xuat_kho_tab(self):
        self.menuBar.setVisible(True)
        self.tabWidget.setCurrentIndex(3) 
        self.lb003.setText("Xuất kho")
        
    def show_ton_kho_tab(self):
        self.menuBar.setVisible(True)
        self.tabWidget.setCurrentIndex(4)
        self.lb003.setText("Tồn kho")
    
    def show_chuyen_vi_tri_tab(self):
        # Chuyển đến tab "Chuyển vị trí"
        self.tabWidget.setCurrentIndex(7)
        self.lb003.setText("Chuyển vị trí cuộn vải")
        self.lb701.setText("")
        self.lb702.setText("")
        self.lb703.setText("")
        self.lb704.setText("")
        self.lb705.setText("")
        self.tb701.setFocus()
    def tong_so_dong(self):
        rows = self.tableWidget.rowCount()
        self.lb101.setText(f"Tổng số dòng dữ liệu: {rows}")
        self.lb101.setStyleSheet("color: rgb(0, 255, 0);")
        
    def tong_so_dong_xa_vai(self):
        rows = self.tableWidget_2.rowCount()
        self.lb201.setText(f"Tổng số dòng dữ liệu: {rows}")
        self.lb201.setStyleSheet("color: rgb(0, 255, 0);")
        
    def tong_so_dong_xuat_kho(self):
        rows = self.tableWidget_5.rowCount()
        self.lb301.setText(f"Tổng số dòng dữ liệu: {rows}")
        self.lb301.setStyleSheet("color: rgb(0, 255, 0);")    
    
    def tong_so_dong_ton_kho(self):
        rows = self.tableWidget_6.rowCount()
        self.lb401.setText(f"Tổng số dòng dữ liệu: {rows}")
        self.lb401.setStyleSheet("color: rgb(0, 255, 0);") 
            
    def delete_row(self, row):
        self.tableWidget.removeRow(row)
        self.tong_so_dong()
        
    def import_from_excel(self):
        # Mở hộp thoại để chọn tệp
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly
        file_to_open, _ = QFileDialog.getOpenFileName(self, "Chọn tệp Excel", "", "Excel Files (*.xlsx *.xls)", options=options)

        # Kiểm tra nếu người dùng không chọn tệp
        if not file_to_open:
            QMessageBox.information(self, "Thông báo", "Không có tệp nào được chọn!")
            return

        # Đọc tệp Excel
        try:
            df = pd.read_excel(file_to_open,sheet_name="NHẬP",usecols=range(9),header=0,skiprows=1)
            df = df.rename(columns={
                'Ngày nhận':'NGAY_NHAN',
                'Style' : 'STYLE',
                'Loại vải' : 'LOAI_VAI',
                'ĐVT' : 'DVT',
                'Lot' : 'LOT',
                'Màu' : 'MAU',
                'Cuộn số' : 'CUON_SO',
                'Số yard' : 'SO_YARD'
            })
            df['NGAY_NHAN'] = pd.to_datetime(df['NGAY_NHAN'], errors='coerce').dt.strftime('%Y-%m-%d')
            df['NHA_MAY'] = self.lb000.text()
            df = df.drop_duplicates(subset=['STYLE','MO','LOAI_VAI','LOT','MAU','CUON_SO'])
            
            dtype = {
                'NGAY_NHAN' : DATE,
                'STYLE' : VARCHAR(20),
                'MO' : VARCHAR(30),
                'LOAI_VAI' : VARCHAR(20),
                'DVT' : VARCHAR(10),
                'LOT' : VARCHAR(50),
                'MAU': VARCHAR(20),
                'CUON_SO': INTEGER,
                'SO_YARD': DECIMAL(6,2),
                'NHA_MAY': VARCHAR(5)
            }
            import_to_sql(df,'DANH_SACH_CUON_VAI',dtype,engine_1)
            QMessageBox.information(self, "Thông báo", f"Tải lên thành công {df.shape[0]} dòng dữ liệu!")
            self.search_nhap_kho()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi đọc tệp Excel: {e}")    
    
    def delete_selected_rows(self):
        # Lấy ID danh sách các hàng được chọn
        selected_IDs  = set(self.tableWidget.item(index.row(),10).text() 
                          for index in self.tableWidget.selectedIndexes()
                          if self.tableWidget.item(index.row(),10) is not None)
        # Kiểm tra nếu không có hàng nào được chọn
        if not selected_IDs:  # Tập hợp rỗng
            QMessageBox.information(self, "Thông báo", "Chưa có dòng nào được chọn")
            return
        # Hiển thị cảnh báo xác nhận
        reply = QMessageBox.question(
            self,
            "Xác nhận xóa",
            f"Bạn có chắc chắn muốn xóa {len(selected_IDs)} dòng đã chọn không?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return  # Không làm gì nếu người dùng chọn "No"
        try:
            connection = connect_to_db()
            if connection is None:
                self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
                return
            cursor = connection.cursor()
            # Chuyển danh sách ID thành chuỗi tham số
            placeholders = ", ".join(["?"] * len(selected_IDs))
            query = f"DELETE FROM DANH_SACH_CUON_VAI WHERE ID IN ({placeholders})"
            # Thực thi câu lệnh với tham số
            cursor.execute(query, tuple(selected_IDs))
            connection.commit()
            
            QMessageBox.information(self, "Thông báo", f"Xóa thành công {len(selected_IDs)} dữ liệu!")
            self.search_nhap_kho()
        except Exception  as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi xóa dữ liệu: {e}")    
        finally:
            if connection:
                connection.close()   
                
    def tai_xuong_file_mau(self):
        headers = ['Ngày nhận','Style','MO','Loại vải','ĐVT','Lot','Màu','Cuộn số','Số yard','Vị trí','ID','Thời gian xả vải','Thời gian xuất kho','Trạng thái']
        df = table_to_dataframe(self.tableWidget,headers)
        # df = df.drop(columns=['ID','Thời gian xả vải','Thời gian xuất kho','Trạng thái'])
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "NHẬP"

        # Add the title to A1
        sheet['A1'] = "MẪU FILE NHẬP VÀO HỆ THỐNG"
        # Merge cells from A1 to J1
        sheet.merge_cells('A1:I1')

        # Center-align the merged cells
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        sheet['A1'].font = Font(color="FFFFFF", bold=True,size=14)
        sheet['A1'].fill = PatternFill(start_color="349eeb", end_color="349eeb", fill_type="solid")
        # Apply bold style to headers (A2:J2)
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=14):
            for cell in row:
                cell.font = Font(bold=True)
                if cell.column <=9:
                    cell.font = Font(color="FF0000",bold=True)
        # Write DataFrame to the sheet, starting from the second row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
   
        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_name = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Nhập kho vải {file_name}", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")
    
    def tai_xuong_file_xa_vai(self):
        headers = ['Ngày nhận','Style','MO','Loại vải','ĐVT','Lot','Màu','Cuộn số','Số yard','Vị trí','ID','Thời gian xả vải','Số giờ xả']
        df = table_to_dataframe(self.tableWidget_2,headers)
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "XẢ VẢI"

        # Add the title to A1
        sheet['A1'] = "DỮ LIỆU XẢ VẢI"
        # Merge cells from A1 to J
        sheet.merge_cells('A1:M1')
        # Center-align the merged cells
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        sheet['A1'].font = Font(color="FFFFFF", bold=True,size=14)
        sheet['A1'].fill = PatternFill(start_color="349eeb", end_color="349eeb", fill_type="solid")
        # Apply bold style to headers (A2:J2)
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=13):
            for cell in row:
                cell.font = Font(bold=True)
        # Write DataFrame to the sheet, starting from the second row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
   
        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_name = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Xả vải {file_name}", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")
            
    def tai_xuong_file_xuat_kho(self):
        headers = ['Ngày nhận','Style','MO','Loại vải','ĐVT','Lot','Màu','Cuộn số','Số yard','Vị trí','ID','Thời gian xả vải','Thời gian xuất kho']
        df = table_to_dataframe(self.tableWidget_5,headers)
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "XUẤT KHO"

        # Add the title to A1
        sheet['A1'] = "DỮ LIỆU XUẤT KHO"
        # Merge cells from A1 to J
        sheet.merge_cells('A1:M1')

        # Center-align the merged cells
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        sheet['A1'].font = Font(color="FFFFFF", bold=True,size=14)
        sheet['A1'].fill = PatternFill(start_color="349eeb", end_color="349eeb", fill_type="solid")
        # Apply bold style to headers (A2:J2)
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=13):
            for cell in row:
                cell.font = Font(bold=True)
        # Write DataFrame to the sheet, starting from the second row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
   
        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_name = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Xuất kho {file_name}", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")
    def tai_xuong_file_ton_kho(self):
        headers = ['Ngày nhận','Style','MO','Loại vải','ĐVT','Lot','Màu','Cuộn số','Số yard','Vị trí','ID']
        df = table_to_dataframe(self.tableWidget_6,headers)
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TỒN KHO"

        # Add the title to A1
        sheet['A1'] = "DỮ LIỆU TỒN KHO"
        # Merge cells from A1 to J
        sheet.merge_cells('A1:K1')

        # Center-align the merged cells
        sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
        sheet['A1'].font = Font(color="FFFFFF", bold=True,size=14)
        sheet['A1'].fill = PatternFill(start_color="349eeb", end_color="349eeb", fill_type="solid")
        # Apply bold style to headers (A2:J2)
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1, max_col=13):
            for cell in row:
                cell.font = Font(bold=True)
        # Write DataFrame to the sheet, starting from the second row
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
   
        # Open file dialog to select save location
        options = QFileDialog.Options()
        file_name = datetime.now().strftime("%d-%m-%Y %H-%M-%S")
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Tải xuống file", 
            f"Tồn kho {file_name}", 
            "Excel Files (*.xlsx);;All Files (*)", 
            options=options
        )

        if not file_path:
            # QMessageBox.information(self, "Thông báo", "Bạn đã hủy việc tải xuống.")
            return

        # Save the file to the selected location
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Thông báo", f"File đã được tải xuống thành công tại:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi lưu file: {e}")
    
    def print_labels_3x2(self):
         # Lặp qua các dòng được chọn để lấy dữ liệu
        selected_rows = self.tableWidget.selectedIndexes()
        if not selected_rows:
            QMessageBox.information(self, "Thông báo", "Chưa có dòng nào được chọn")
            return
        
        # Tạo danh sách các hàng (độc nhất) được chọn
        rows = set(index.row() for index in selected_rows)
        
        # Hiển thị cảnh báo xác nhận
        reply = QMessageBox.question(
            self,
            "Xác nhận in tem",
            f"Bạn có chắc chắn muốn in tem {len(rows)} dòng đã chọn không?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        if reply == QMessageBox.No:
            return  # Không làm gì nếu người dùng chọn "No"

        # Lấy dữ liệu từ các cột cần thiết
        data_list = []
        for row in rows:
            data = {
                "ID": self.tableWidget.item(row, 10).text(),
                "NgayNhap": self.tableWidget.item(row, 0).text(),
                "Style": self.tableWidget.item(row, 1).text(),
                "MO": self.tableWidget.item(row, 2).text(),
                "LoaiVai": self.tableWidget.item(row, 3).text(),
                "Dvt": self.tableWidget.item(row, 4).text(),
                "Lot": self.tableWidget.item(row, 5).text(),
                "Mau": self.tableWidget.item(row, 6).text(),
                "CuonSo": self.tableWidget.item(row, 7).text(),
                "SoYard": self.tableWidget.item(row, 8).text(),
            }
            data_list.append(data)

        try:
            # Tạo file PDF tạm thời
            temp_dir = tempfile.gettempdir()
            output_path = os.path.join(temp_dir, "labels_3x2.pdf")

            # Định kích thước trang 4x6 inch
            page_width = 3.1 * inch
            page_height = 2 * inch

            # Tạo PDF
            c = canvas.Canvas(output_path, pagesize=(page_width, page_height))

            for data in data_list:
                # Reset vị trí in (trang mới cho mỗi tem)
                gap = 10
                # In ID (in đậm, kích thước 20)
                c.setFont("Arial-Bold", 20)  # Font in đậm
                c.drawString(15, page_height - gap*3, f"ID: {data['ID']}")
                # In các thông tin khác (kích thước chữ 12)
                c.setFont("Arial", 9)  # Font thường
                c.drawString(15, page_height - gap*5, f"Ngày nhập: {data['NgayNhap']}")
                c.drawString(15, page_height - gap*6, f"Style: {data['Style']}")
                c.drawString(15, page_height - gap*7, f"MO: {data['MO']}")
                c.drawString(15, page_height - gap*8, f"Loại vải: {data['LoaiVai']}")
                c.drawString(15, page_height - gap*9, f"ĐVT: {data['Dvt']}")
                c.drawString(15, page_height - gap*10, f"Lot: {data['Lot']}")
                c.drawString(15, page_height - gap*11, f"Màu: {data['Mau']}")
                c.drawString(15, page_height - gap*12, f"Cuộn số: {data['CuonSo']}")
                c.drawString(15, page_height - gap*13, f"Số yard: {data['SoYard']}")

                # Tạo mã QR từ ID
                qr = qrcode.make(data['ID'])
                qr_temp_path = os.path.join(temp_dir, f"qr_{data['ID']}.png")
                qr.save(qr_temp_path)

                # Vẽ mã QR lên tem
                qr_size = 0.8 * inch  # Kích thước mã QR 2x2 inch
                c.drawImage(qr_temp_path, page_width - qr_size - 5, page_height - qr_size - 5, width=qr_size, height=qr_size)
                c.drawImage(qr_temp_path, page_width - qr_size - 5, 5, width=qr_size, height=qr_size)
                # Đường dẫn resource trong file .qrc
                logo_resource_path = ":/pictures/pics/canh_bao.png"
                # Tạo đường dẫn tạm thời cho ảnh
                temp_dir = tempfile.gettempdir()
                temp_logo_path = os.path.join(temp_dir, "canh_bao.png")
                # Lưu ảnh từ resource ra file tạm
                pixmap = QPixmap(logo_resource_path)
                if not pixmap.isNull():
                    pixmap.save(temp_logo_path, "PNG")
                else:
                    raise FileNotFoundError(f"Không thể tìm thấy resource: {logo_resource_path}")
                # Vẽ mã tem cảnh báo lên tem
                # logo_size = 1.5 * inch  # Kích thước mã QR 2x2 inch
                # c.drawImage(temp_logo_path, 10,10, width=logo_size, height=logo_size)
                # Lưu mỗi tem trên một trang
                c.showPage()

                # Xóa file QR tạm
                if os.path.exists(qr_temp_path):
                    os.remove(qr_temp_path)

            c.save()
            # # Gửi lệnh in trực tiếp tới máy in
            printer_name = win32print.GetDefaultPrinter()
            win32api.ShellExecute(
                0,
                "print",
                output_path,
                None,
                ".",
                0
            )

            QMessageBox.information(self, "Thông báo", "In tem thành công!")

        except Exception as e:
            QMessageBox.critical(self, "Lỗi", f"Đã xảy ra lỗi khi in tem: {e}")
            
    def handle_scan_xa_vai(self):
        qr_code = self.tb501.text().strip()

        if not qr_code:
            return

        # Kiểm tra QR trong database
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        query = "SELECT ID FROM DANH_SACH_CUON_VAI WHERE ID = ? and TRANG_THAI = N'Nhập kho'"
        cursor.execute(query, (qr_code,))
        result = cursor.fetchone()

        if result:
            # Nếu QR hợp lệ, cập nhật THOI_DIEM_XUAT_KHO
            update_query = """
                UPDATE DANH_SACH_CUON_VAI 
                SET THOI_GIAN_XA = ? ,
                TRANG_THAI = N'Xả vải',
                VI_TRI = ''
                WHERE ID = ?
            """
            cursor.execute(update_query, (datetime.now(), qr_code))
            connection.commit()

            # Hiển thị thông báo thành công
            self.lb501.setText("OK")
            self.lb502.setText("")
            self.lb503.setText(f"ID : {qr_code}")
            self.lb504.setText(f"Thời điểm xả vải : {datetime.now()}")
            QSound.play(":/sounds/sounds/bellding.wav") # Phát âm thanh thành công
        else:
            # Nếu QR không hợp lệ
            self.lb501.setText("")
            self.lb502.setText("Mã QR không hợp lệ!")
            self.lb503.setText(f"ID : {qr_code}")
            self.lb504.setText("")
            QSound.play(":/sounds/sounds/error.wav") # Phát âm thanh lỗi
            
        # Reset lại LineEdit
        self.tb501.clear()
        self.tb501.setFocus()
    def handle_scan_xuat_kho(self):
        qr_code = self.tb601.text().strip()

        if not qr_code:
            return

        # Kiểm tra QR trong database
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        query = "SELECT ID FROM DANH_SACH_CUON_VAI WHERE ID = ? AND TRANG_THAI IN (N'Nhập kho',N'Xả vải')"
        cursor.execute(query, (qr_code,))
        result = cursor.fetchone()

        if result:
            # Nếu QR hợp lệ, cập nhật THOI_DIEM_XUAT_KHO
            update_query = """
                UPDATE DANH_SACH_CUON_VAI 
                SET THOI_GIAN_XUAT_KHO = ? ,
                TRANG_THAI = N'Xuất kho',
                VI_TRI = ''
                WHERE ID = ?
            """
            cursor.execute(update_query, (datetime.now(), qr_code))
            connection.commit()

            # Hiển thị thông báo thành công
            self.lb601.setText("OK")
            self.lb602.setText("")
            self.lb603.setText(f"ID : {qr_code}")
            self.lb604.setText(f"Thời điểm xuất kho : {datetime.now()}")
            QSound.play(":/sounds/sounds/bellding.wav") # Phát âm thanh thành công
        else:
            # Nếu QR không hợp lệ
            self.lb601.setText("")
            self.lb602.setText("Mã QR không hợp lệ!")
            self.lb603.setText(f"ID : {qr_code}")
            self.lb604.setText("")
            QSound.play(":/sounds/sounds/error.wav") # Phát âm thanh lỗi
            
        # Reset lại LineEdit
        self.tb601.clear()
        self.tb601.setFocus()    
    
    def handle_scan_chuyen_vi_tri(self):
        qr_code = self.tb701.text().strip()

        if not qr_code:
            return

        if qr_code[0:1] in ['A','B','C','D','E','F','G','H','I','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
            self.tb701.clear()
            self.tb701.setFocus() 
            self.lb702.setText("")
            self.lb703.setText(f"{qr_code}")
            self.lb704.setText("")
            self.lb705.setText("")
            QSound.play(":/sounds/sounds/success.wav") # Phát âm thanh thành công
            return
        
        # Kiểm tra QR trong database
        connection = connect_to_db()
        if connection is None:
            self.lb003.setText("Không thể kết nối tới cơ sở dữ liệu!")
            return
        cursor = connection.cursor()
        query = "SELECT ID FROM DANH_SACH_CUON_VAI WHERE ID = ? AND TRANG_THAI IN (N'Nhập kho',N'Xả vải')"
        cursor.execute(query, (qr_code,))
        result = cursor.fetchone()

        if result:
            # Nếu QR hợp lệ, cập nhật THOI_DIEM_XUAT_KHO
            update_query = """
                UPDATE DANH_SACH_CUON_VAI 
                SET VI_TRI = ? 
                WHERE ID = ?
            """
            vi_tri = self.lb703.text().strip()
            if vi_tri:
                cursor.execute(update_query, (vi_tri, qr_code))
                connection.commit()

                # Hiển thị thông báo thành công
                self.lb701.setText("OK")
                self.lb702.setText("")
                self.lb704.setText(f"Vị trí : {vi_tri}")
                self.lb705.setText(f"ID : {qr_code}")
                QSound.play(":/sounds/sounds/bellding.wav") # Phát âm thanh thành công
            else:
                # Nếu chưa cập nhật vị trí hiện tại
                self.lb701.setText("")
                self.lb702.setText("Vui lòng quét mã QR vị trí trước khi quét mã cuộn vải!")
                self.lb704.setText("")
                self.lb705.setText("")
                QSound.play(":/sounds/sounds/error.wav") # Phát âm thanh lỗi
        else:
            # Nếu QR không hợp lệ
            self.lb701.setText("")
            self.lb702.setText("Mã QR không hợp lệ!")
            self.lb704.setText("")
            self.lb705.setText("")
            QSound.play(":/sounds/sounds/error.wav") # Phát âm thanh lỗi
            
        # Reset lại LineEdit
        self.tb701.clear()
        self.tb701.setFocus()    

def main():
    app = QApplication(sys.argv)
    window = MainApp()
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window.show()
    app.exec_()

if __name__ == '__main__':
    main()

